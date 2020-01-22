# -*- coding:UTF-8 -*-
from json.encoder import JSONEncoder

from openpyxl import load_workbook
from macpath import split
import re
from urllib3.fields import RequestField
from urllib3.filepost import encode_multipart_formdata
from coingame.config.globalparam import get_testEv_ip_name_pwd
from public.common.log import Log
from decimal import Decimal

import paramiko
import os,sys
import smtplib
from email.mime.text import MIMEText
from email.header import Header
import datetime


logger = Log()

def list_to_may(key,ls):
    result_map = {}
    for l in ls:
        result_map.setdefault(l[key],l)
    return result_map
    # print (result_map)


def truncate(f, n):
    '''Truncates/pads a float f to n decimal places without rounding'''
    #当你不再需要该表时， 用 drop；当你仍要保留该表，但要删除所有记录时， 用 truncate；
    # 当你要删除部分记录时（always with a WHERE clause), 用 delete.
    s = '{}'.format(f)
    if 'e' in s or 'E' in s:
        return '{0:.{1}f}'.format(f, n)
    i, p, d = s.partition('.')
    v =  '.'.join([i, (d+'0'*n)[:n]])
    return float(v)

def create_up_file_data(file,name='file',content_type='image/png',boundary=''):
    """创建上传文件请求数据
    注：代码来源于request.py （源码中无法指定 boundary）
    """
    file_data = open(file,'rb').read()
    filename = re.search('.*/(.*)',file).group(1)
    new_fields = []
    rf = RequestField(name=name,data=file_data,filename=filename)
    rf.make_multipart(content_type=content_type)
    new_fields.append(rf)
    body, content_type = encode_multipart_formdata(new_fields,boundary=boundary)
    return body



def add_sub_date(old_date,day:int=None,date_format='%Y-%m-%d %H:%M:%S'):
    """加减日期"""
    final_date = None
    if day != None:
        delta=datetime.timedelta(days=day)  #时间差
        final_date = old_date+delta
    return final_date.strftime(date_format)

def add_sec_date(old_date,seconds:int=None,date_format='%Y-%m-%d %H:%M:%S'):
    """加减秒"""
    final_date = None
    if seconds != None:
        delta=datetime.timedelta(seconds=seconds)
        final_date = old_date+delta
    return final_date.strftime(date_format)

def read_write_info(file_path,info=None):
    """向文件写或读取信息"""
    if info == None:
        fp = open(file_path,'r',encoding='utf8')
        # print(fp.read())
        return fp.read()
    else:
        fp = open(file_path,'w',encoding='utf8')
        fp.write(info)
    fp.close()

def send_email(sender='lusainan13579@163.com',login_pwd='150061803670513',
               receivers=['lusainan13579@163.com']
               ,subject='Python SMTP 邮件测试',
               msg='Python 邮件发送测试...'):
    """发送邮件"""
    message = MIMEText(msg, 'plain', 'utf-8')
    # message['From'] = Header("菜鸟教程", 'utf-8')   # 发送者
    # message['To'] =  Header("测试", 'utf-8')        # 接收者
    message['Subject'] = Header(subject, 'utf-8')

    try:
        smtpObj = smtplib.SMTP_SSL('smtp.163.com') #SMTP服务器
        smtpObj.login(sender,login_pwd)  #不是登录邮箱密码，而是授权码
        smtpObj.sendmail(sender, receivers, message.as_string())
        smtpObj.quit()
        logger.info ("邮件发送成功")
    except smtplib.SMTPException:
        logger.info ("Error: 无法发送邮件")


def del_file(path,is_delete_all=False):
    """删除指定目录下文件"""
    ls = os.listdir(path)  #列出目录下所有文件,[],list
    # print("ls:",ls)  #ls: ['file1.txt', 'file2.txt', 'file3.txt']
    for i in ls:
        c_path = os.path.join(path, i) #获取目录下文件的路径
        print(c_path)  #D:\auto_test_nancy\coingame\data\temp\file1.txt
        if os.path.isdir(c_path):      # True 删除文件夹夹下文件
            if is_delete_all == True:
                del_file(c_path)   #删除目录下文件
        else:
            os.remove(c_path)      #删除文件  目录不能os.remove() 删除文件

def get_all_files_in_local_dir(local_dir):
    # 保存所有文件的列表   只支持local_dir下全是file的情况！
    all_files = []
    # 获取当前指定目录下的所有目录及文件，包含属性值
    files = os.listdir(local_dir)
    # print(files)  #['dir1', 'dir2', 'file1.txt', 'file2.txt', 'file3.txt']
    for x in files:
        # local_dir目录中每一个文件或目录的完整路径
        filename = os.path.join(local_dir, x)
        # print(filename)
        # # 如果是目录，则递归处理该目录
        # if os.path.isdir(filename):
        #     all_files.extend(get_all_files_in_local_dir(filename))
        # else:
        #     all_files.append(filename)

        if os.path.isdir(filename): #目录就中断
            break
        all_files.append(filename)  #将所有文件路径拼成一个list
    return all_files

class MyObject ():

    def __init__(self,json_dict):
        self.dict_to_obj(json_dict)
        pass

    def dict_to_obj(self,d): #字典转对象
        seqs = tuple, list, set, frozenset
        for i, j in d.items():
            if isinstance(j, dict):
                setattr(self, i, dict_to_obj(j))
            elif isinstance(j, seqs):
                setattr(self, i,
                    type(j)(dict_to_obj(sj) if isinstance(sj, dict) else sj for sj in j))
            else:
                setattr(self, i, j)
        return self

    def __repr__(self):
        #并没有没外部调用??阿略略
        for k,v in self.__dict__.items():
            if isinstance(v,type): #isinstance()判断的是一个对象是否是该类型本身，或者位于该类型的父继承链上
                # print(k)
                if 'public.common.util.MyObject' in str(v):
                    print('测试',v.__dict__)
                    # self.__dict__[]
        #     else:
        #         print('其它',k,type(v))
        # return repr(self.__dict__)


#把字典转换成复杂的对象(传入参数为字典)
def dict_to_obj(d):
    top = type('MyObject', (object,), d)
    seqs = tuple, list, set, frozenset
    for i, j in d.items():
        try:
            if isinstance(j, dict):
                setattr(top, i, dict_to_obj(j))
            elif isinstance(j, seqs):
                setattr(top, i,
                    type(j)(dict_to_obj(sj) if isinstance(sj, dict) else sj for sj in j))
            else:
                setattr(top, i, j) #setattr(object, name, value) 对象  对象属性 属性值
        except:
            pass
    return top



def getMapValue(m,key):
    """获取字典中的值 如果没有key 则返回None"""
    try:
        # print(m[key])
        return m[key]
    except:
        # print(None)
        return None

def setVarsValue(varsName,value,params):
    """替换掉${varsName}中的值  参照jmeter${varsName}"""
    return re.sub(r'\D', value, params)#移除非数字
    # return re.sub('\$\{%s\}'%varsName, value, params)

def twoListToMap(listKey,listValue):
    """把2个列表转换成字典"""
    returnMap = {}
    for i  in range(len(listKey)):
        returnMap[listKey[i]]=listValue[i]
    return returnMap

def calculate_float_data(first_data,second_data,calculate_type='-'):
    # up_int = 100000000000
    # # first_data = first_data * up_int
    # # second_data = second_data * up_int
    # if calculate_type == '-':
    #     return (first_data-second_data)*up_int/(up_int)
    #
    # if calculate_type == '+':
    #     return (first_data+second_data)*up_int/(up_int)
    #
    # if calculate_type == '*':
    #     return (first_data*second_data)*up_int/(up_int)
    #
    # if calculate_type == '/':
    #     return first_data/second_data

    first_data = str(first_data)
    second_data = str(second_data)
    if calculate_type == '-':
        return float(Decimal(first_data)-Decimal(second_data))
    if calculate_type == '+':
        return float((Decimal(first_data)+Decimal(second_data)))

    if calculate_type == '*':
        return float((Decimal(first_data)*Decimal(second_data)))

    if calculate_type == '/':
        return float(Decimal(first_data)/Decimal(second_data))

def search_str(reg,source_str,group_index=0):
    """查找字符串"""
    try:
        #正则匹配 扫描整个字符串并返回第一个成功的匹配。reg正则表达式
        return re.search(reg,source_str).group(group_index)
    except:
        return ''

def sub_str(reg,source_str):
    """截取字符串"""
    try:
        return re.search(reg,source_str).group(1)
    except:
        return ''

class OperatorExcel(object):
    """数据库类"""
    #openpyxl也只能处理xlsx、xlsm文件，对xls也只能改xlwt 、wlrd这个两个模块处理！！！！！！！
    def __init__(self, filePath):    #构造方法
        self.filePath=filePath
        self.wb = load_workbook(filePath)  #openpyxl只支持xlsx格式，老版的xls格式需要其他方法去加载。
        # #获取全部表名
        # print("sheetnames",self.wb.sheetnames)
        #修改表名
        # self.sheet1=self.wb.get_sheet_by_name('test_sheet')
        # self.sheet1=self.wb[self.wb.sheetnames[1]]
        # self.sheet1.title = "caseInfo"
        #获取第一个表名
        # print("sheetname1:",self.wb.sheetnames[0])
        #删除表
        # self.wb.remove(self.wb[self.wb.sheetnames[3]])
        # self.wb.remove(self.wb.get_sheet_by_name(name='Sheet2'))
        # self.wb.remove_sheet(self.wb.get_sheet_by_name(name='new sheet'))
        # print("sheetnames",self.wb.sheetnames)
        # #新建表
        # self.wb.create_sheet(index=0, title="mysheet1")
        # print("sheetnames",self.wb.sheetnames)
        # self.wb.save(filePath) #保存

    def close(self):
        self.wb.close()
            
    def __parseSql(self,sql):  ##私有变量  __XX私有变量、_X实例变量、__XX__特殊变量
        arr=sql.split('&')  #将字符串$sql 按照 & 分割
        self.sheetName=arr[0]      #caseInfo&isRun=true&0  caseInfo表名 len(arr)==3
        self.where = arr[1].split(",");
        # print(self.where)
        if(len(arr)>3):
            self.needField=arr[3]           #needField=自定义的arr[3]  sql变量第四个字段,&后
        self.isNeedOne=True if arr[2]=='1' else False  ##if arr[2]=='1', isNeedOne=True else  False
     
   
    def getData2DictList(self,sql):
        """
         * 该方法支持多条件查询 sheetName&fieldName1=findValue,fieldName2=findValue2,...
         * (或者fieldName<>findValue,...)&inNeedOne
        """
        """1、用法一 读取匹配的行+列，组成[{},{},{}]嵌套，方便后面读取"""
        logger.info('执行的查询语句：'+sql)
        result = []  
        rows = self.__matchCols(sql)  #调用下面__matchCols函数
        # print("1",rows)  #匹配isRun=true的行[2,3,6]
        sheet = self.wb[self.sheetName]
        # print("2",sheet)#<Worksheet "caseInfo">
        colNum = sheet.max_column
        # print("3",colNum)  #列max
        for i in rows:  #匹配几行循环几次  i 行
            temp = {}
            for m in range(1,colNum+1):  #1,2,3,4,5,6,7,8  循环8次 [1,colNum+1)   m列
                value = sheet.cell(row=i,column=m).value
                if value==None:
                    value=''
                temp[sheet.cell(row=1,column=m).value]=value  #temp['key']=value  字典赋值
            # print(temp)  #每个匹配的行组成一个字典
            result.append(temp)  #将字典拼接成list   [{},{},{}]
        self.__clearSql()          #初始化值清空为''
        # print(result)   #拼接好的list  [{},{},{},{}]
        return result

    def getOneField2List(self,sql):
        """
         * 该方法支持多条件查询 sheetName&fieldName1=findValue,fieldName2=findValue2,...
         * (或者fieldName<>findValue,...)&inNeedOne&needField
         * 注:needField为要读取的字段名称!!
        """
        """2、用法二 读取匹配范围内某字段值，组成list[]，方便后面读取"""
        result = []  
        rows = self.__matchCols(sql) #匹配的行
        # print("match rows:",rows) #match rows: [2, 3, 6]
        sheet = self.wb.get_sheet_by_name(self.sheetName)
        fieldMap = self.__fieldAndIndex2Dict(self.sheetName)  #调用下面的私有函数__fieldAndIndex2Dict
        colNum = sheet.max_column
        # print("colNum:",colNum) #8列
        for i in rows:
            value = sheet.cell(row=i,column=fieldMap[self.needField]).value
            # print("value:",value)
            if value==None:
                value=''
#             result.append(sheet.cell(row=i,column=fieldMap[self.needField]).value)
            result.append(value)
        self.__clearSql()
        return result
    
    
    def insertValues(self,dict1,sheetName):
        #插入数据   
        
        sheet = self.wb.get_sheet_by_name(sheetName)
        # print("sheet:",sheet)
        rowNum=sheet.max_row
        # print(rowNum)
        fieldMap = self.__fieldAndIndex2Dict(sheetName)
        # print("字段从1开始排序：",fieldMap)
        for d,v in dict1.items(): 
            try:
                sheet.cell(row=rowNum+1,column=fieldMap[d]).value=v
            except KeyError:
                pass
        self.__clearSql()
        self.__save()
        
    def updateValues(self,dict1,sql):
        #按sql 更新数据   比如这里替换2、3、6行数据
        """
         * 该方法支持多条件查询 sheetName&fieldName1=findValue,fieldName2=findValue2,...
         * (或者fieldName<>findValue,...)&inNeedOne
        """  
        
        rows = self.__matchCols(sql)
        # print(rows)
        sheet = self.wb.get_sheet_by_name(self.sheetName)
        # print(sheet)
        rowNum=sheet.max_row
        # print(rowNum)
        fieldMap = self.__fieldAndIndex2Dict(self.sheetName)
        # print("字段从1开始排序:",fieldMap)  #fieldMap: {'id': 1, 'b_u_id': 2, 'user_name': 3, 'user_password': 4, 'betting_currency_id': 5, 'isRun': 6, 'caseInfo': 7, 'index': 8}
        for i in rows:           
            for d,v in dict1.items(): 
                try:
                    sheet.cell(row=i,column=fieldMap[d]).value=v
                except KeyError:
                    pass 
        self.__clearSql()         
        self.__save()
      
    def __save(self):
        # print("self.filePath:",self.filePath)
        self.wb.save(self.filePath)

        
      
    def __matchCols(self,sql):
        #匹配满足条件的行数   
        self.__parseSql(sql) #调用上面的__parseSql(sql) #获取self.sheetName、self.where、self.isNeedOne
        sheet = self.wb[self.sheetName]
        fieldMap = self.__fieldAndIndex2Dict(self.sheetName)  #调用下面的__fieldAndIndex2Dict
        rowNum = sheet.max_row
        colNum = sheet.max_column
        result = []
        for i in range(2,rowNum+1):
            for j in range(len(self.where)):
                where= re.split('=|<>',self.where[j])
                findField = where[0]
                findValue = where[1]
                value = sheet.cell(row=i,column=fieldMap[findField]).value
                if not self.__matchFindType(self.where[j], value, findValue):
                    break
            #此处的else（如果前面语句没有发生break语句 则会执行else 否则不执行else）
            else:
                result.append(i)   
                #如果只需要一条数据则返回 否则继续
                if self.isNeedOne:
                    break
        return result
        # 1 [2, 3，6] 满足sql查询条件的行数

    #怎么理解?
    def __matchFindType(self,where,haveFindValue,exceptValue):
        findType='equals'
        haveFindValue=str(haveFindValue)
        if where.__contains__('<>'):
            findType="except"
        if findType=='equals':
            if exceptValue==haveFindValue:
                return True
        elif findType=='except':
            if exceptValue != haveFindValue:
                return True
        return False
  
      
      
    #把第一行的值 和 列号存入字典中  key=field value=列号（从1开始）
    #{'id': 1, 'b_u_id': 2, 'user_name': 3, 'user_password': 4, 'betting_currency_id': 5, 'isRun': 6, 'caseInfo': 7, 'index': 8}
    def __fieldAndIndex2Dict(self,sheetName):
        sheet = self.wb[sheetName]
        colNum=sheet.max_column
        map = {}
        for i in range(colNum):
            key = sheet.cell(row=1,column=i+1).value
            map[key]=i+1
        return map
     
    def __clearSql(self):
        self.sheetName=''
        self.where=''
        self.isNeedOne=False
        self.needField=''

class SSH_Util():

    def __init__(self,ip='',username='',password='',testEv=''):
        if testEv != '':
            dict = get_testEv_ip_name_pwd(testEv)   #传testEv，从globalparam.py的dict获取默认值
            self.ip = dict['ip']
            self.username = dict['name']
            self.password = dict['pwd']
            # print(testEv,dict)
        else:
            self.ip =ip                            #没传testEv，传ip、username、password
            # self.port = port
            self.username = username
            self.password = password
            # print(None)
            # print({"ip":self.ip,"username":self.username,"password":self.password})

    def rum_cmd(self,cmd):
        logger.info('重启服务：%s'%cmd)
        ssh = paramiko.SSHClient()#创建SSH对象
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())#允许连接不在know_hosts文件中的主机
        ssh.connect(hostname=self.ip,username=self.username, password=self.password)#连接服务器
        # print({"ip":self.ip,"username":self.username,"password":self.password})
        stdin, stdout, stderr = ssh.exec_command(cmd)#执行命令并获取命令结果
        #stdin为输入的命令
        #stdout为命令返回的结果
        #stderr为命令错误时返回的结果
        res,err = stdout.read(),stderr.read()
        result = res if res else err
        ssh.close()#关闭连接
        print(result)
        return result

    def up_or_down_file(self,server_path, local_path,is_up = True):
        #上传/下载单个文件
        transport = paramiko.Transport((self.ip,22))
        transport.connect(username=self.username, password=self.password)
        sftp = paramiko.SFTPClient.from_transport(transport)#如果连接需要密钥，则要加上一个参数，hostkey="密钥"
        if is_up:
            sftp.put(local_path, server_path)#将本地的Windows.txt文件上传至服务器/root/Windows.txt
        else:
            sftp.get(server_path, local_path)#将Linux上的/root/Linux.txt下载到本地
        transport.close()#关闭连接

        
if __name__=='__main__':
    # list_to_may()  #???
    # truncate()   #???
    # file='D:\\auto_test_nancy\\coingame\\data\\file1.txt'
    # print(create_up_file_data(file)) #???



    #now=datetime.datetime.now()
    # print("now:",now)
    # print(add_sub_date(now,1)) #明天
    # print(add_sub_date(now,-1)) #昨天
    # print(add_sec_date(now,600)) #10min后
    # print(add_sec_date(now,-600)) #10min前

    print(read_write_info("D:\\auto_test_nancy\\coingame\\data\\file1.txt"))#读取
    # read_write_info("D:\\auto_test_nancy\\coingame\\data\\file1.txt","你好呀1")#写入
    # send_email()
    # del_file("D:\\auto_test_nancy\\coingame\\data\\temp\\",False) #False只删除目录下一级file，True删除目录下所有层级file
    # get_all_files_in_local_dir("D:\\auto_test_nancy\\coingame\\data\\temp\\dir1")

    #####o=MyObject({'key1':'1','key2':'2'})
    # o.dict_to_obj({'key1':'1','key2':'2'})

    #print(dict_to_obj({'key1':'1','key2':'2','key3':'3','key4':'4'})) #<class '__main__.MyObject'>  不理解??
    # print(getMapValue({'key1':'1','key2':'2','key3':'3','key4':'4'},"key1")) #根据key，获取对应value OK
    # print(getMapValue({'key1':'1','key2':'2','key3':'3','key4':'4'},"key5")) #找不到key，返回None OK
    # print(setVarsValue("-","","2004-959-559"))#OK 替换掉字符串中啥
    # print(twoListToMap(['name','age','sex'],['tom','18','man'])) #{'name': 'tom', 'age': '18', 'sex': 'man'} OK
    # print(calculate_float_data('12.2','1.91')) #十进制计算  高精度  OK
    # print(search_str('(\d+).*?(\d+).*','Hello 123456789 Word_This is just a test 666 Test')) #OK
    # print(sub_str('(\d+).*?(\d+).*','Hello 123456789 Word_This is just a test 666 Test'))    #OK

    # y= SSH_Util('','','','aa')  #取默认
    ##############y= SSH_Util('192.168.100.118','root','lusainan','') #自定义
    # y.up_or_down_file("/opt/temp/dir1/2.png",\
    #                   "D:/web_http/112244/999.png",is_up = True)   #is_up【上传】  下载

    # y.up_or_down_file("/opt/temp/dir1/2.png",\
    #                   "D:/web_http/112244/888.png",is_up = False)    #is_up上传  【下载】
    # y.rum_cmd("cd /opt/temp/dns-service/build/libs;ps -ef|grep dns-service-1.0.2-RELEASE.jar")
    # y.rum_cmd("cd /opt/temp/dns-service/build/libs;java -jar dns-service-1.0.2-RELEASE.jar &")#NG
    ####### y.rum_cmd("cd /opt/temp;pwd")

    # path = 'D:\\auto_test_nancy\\coingame\\data\\cases.xlsx'
    # excel = OperatorExcel(path)
    # data = excel.getData2DictList('caseInfo&isRun=true&0')  #将表caseInfo，字段值isRun=true匹配的行列数据组成list，每一行是一个dict  [{},{},{}]
    # data=excel.getOneField2List('caseInfo&isRun=true&0&caseInfo') #将表caseInfo,字段值isRun=true匹配行,且将字段caseInfo的字段值取出，拼接成list  [a,b,c]  0所有匹配(非1)  1匹第一个  if arr[2]=='1', isNeedOne=True else  False
    # data=excel.insertValues({'id':'12', 'b_u_id':'UB1','user_name':'AUTO08@qq.com'},"caseInfo") #末尾加一行
    #data=excel.updateValues({'id': '1', 'b_u_id': 'UB1', 'user_name': 'AUTO21@qq.com'},"caseInfo&isRun=true&0") #修改所有匹配
    # print(data)
    # print(type(data))

    pass