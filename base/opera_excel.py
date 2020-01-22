# -*- coding: utf-8 -*-
# __author__ = 'lusn'
import requests
import json
import xlrd
from openpyxl import load_workbook
import openpyxl


#xlrd读excel数据=========
class ExcelUtil():
    def __init__(self, excelPath, sheetName="Sheet1"):
            self.data = xlrd.open_workbook(excelPath)
            self.table = self.data.sheet_by_name(sheetName)
            # 获取第一行作为key值
            self.keys = self.table.row_values(0)
            # 获取总行数
            self.rowNum = self.table.nrows
            # 获取总列数
            self.colNum = self.table.ncols
    def dict_data(self):
        if self.rowNum <= 1:
            print("总行数小于1")
        else:
            r = []
            j = 1
            for i in list(range(self.rowNum-1)):
                s = {}
                # 从第二行取对应values值
                s['rowNum'] = i+2
                values = self.table.row_values(j)  #每一行值，表头[0],第一行数据[1]
                for x in list(range(self.colNum)):
                    s[self.keys[x]] = values[x]  #一行一个字典
                r.append(s)                      #N行字典拼接成一个list
                j += 1
            return r

#copy一份数据
def copy_excel(excelpath1, excelpath2):
    '''复制excel，把excelpath1数据复制到excelpath2'''
    wb2 = openpyxl.Workbook()
    wb2.save(excelpath2)
    # 读取数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row         # 最大行数
    max_column = sheet1.max_column   # 最大列数
    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):   # chr(97)='a'
            n = chr(n)                            # ASCII字符
            i ='%s%d'% (n, m)                     # 单元格编号
            cell1 = sheet1[i].value               # 获取data单元格数据
            sheet2[i].value = cell1               # 赋值到test单元格
    wb2.save(excelpath2)                 # 保存数据
    wb1.close()                          # 关闭excel
    wb2.close()


#openpyxl写入数据
class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet
    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)


if __name__ == '__main__':
    #xlrd读excel数据
    filepath = "D:/explore_dfw/data/case/explore_union_case.xlsx"
    sheetName = "Sheet1"
    data = ExcelUtil(filepath,sheetName)
    print(data.dict_data())  #先从excel里面读取测试数据，返回字典格式，组成list

    #openpyxl写入数据
    #copy_excel("D:/explore_dfw/data/case/explore_union_case.xlsx", "D:/explore_dfw/data/case/explore_union_case_copy.xlsx")
    wt = Write_excel("D:/explore_dfw/data/case/explore_union_case_copy.xlsx")
    wt.write(4, 5, "HELLEOP") #行=4  列=5，写入HELLEOP
    wt.write(4, 6, "HELLEOP") #行=4  列=6，写入HELLEOP
