# -*- coding: utf-8 -*-
# __author__ = 'lusn'
import requests
import json
import xlrd
from openpyxl import load_workbook
import openpyxl


class Webrequests():
    #封装POST
    def httpPost(self,url,data,header=None):
        res = None
        if header !=None:
            res=requests.post(url,data=data,headers=header)
        else:
            res = requests.post(url=url,data=data)
        return res

    #封装GET
    def httpGet(self,url,data,header=None):
        res = None
        if header !=None:
            res = requests.get(url=url,data=data,headers=header,verify=False)
        else:
            res = requests.get(url=url,data=data,verify=False)
        return res

    #封装Request
    def httpGetOrPost(self,method,url,data,header):
        if method in "get":
            mres=self.httpGet(url,data,header)
        elif method == "post":
            mres=self.httpPost(url,data,header)
        elif method in"put":
            mres=requests.put(url,data=data,headers=headers)
        elif method in "delete":
            mres=requests.delete(url,data=data,headers=headers)
        else:
            mres = requests.post(url, data=data, headers=headers)
            print("错误")
        return mres.json()

    #case1
    def test0001(self):
        self.url="http://172.17.3.187:8080/auth/login"
        self.method="post"
        self.data='{"user":"admin","password":"%s","network":"east-network"}' %("adminpw")
        self.header={"Content-Type": "application/json"}
        d=self.httpGetOrPost(self.method,self.url,self.data,self.header)
        if d["message"]=="You have successfully logged in!":
            print("登录 PASS")
        else:
            print('exception:',d["message"])


    #封装excel
    def readSheetdata(self,cell):
        wb=load_workbook(r'D:\auto_test_nancy\dfw_union\data\explore_union_case.xlsx')
        sheet=wb.active
        value=sheet[cell]
        print(value.value,type(value.value))
        return value.value

#xlrd读excel数据
class ExcelUtil():
    def __init__(self, excelPath, sheetName="Sheet1"):
            self.data = xlrd.open_workbook(excelPath)
            self.table = self.data.sheet_by_name(sheetName)
            # 获取第一行作为key值
            self.keys = self.table.row_values(0)
            # 获取总行数
            self.rowNum = self.table.nrows
            # 获取总列数
            self.colNum = self.table.ncols
    def dict_data(self):
        if self.rowNum <= 1:
            print("总行数小于1")
        else:
            r = []
            j = 1
            for i in list(range(self.rowNum-1)):
                s = {}
                # 从第二行取对应values值
                s['rowNum'] = i+2
                values = self.table.row_values(j)
                for x in list(range(self.colNum)):
                    s[self.keys[x]] = values[x]
                r.append(s)
                j += 1
            return r

#copy一份数据
def copy_excel(excelpath1, excelpath2):
    '''复制excel，把excelpath1数据复制到excelpath2'''
    wb2 = openpyxl.Workbook()
    wb2.save(excelpath2)
    # 读取数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row         # 最大行数
    max_column = sheet1.max_column   # 最大列数
    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):   # chr(97)='a'
            n = chr(n)                            # ASCII字符
            i ='%s%d'% (n, m)                     # 单元格编号
            cell1 = sheet1[i].value               # 获取data单元格数据
            sheet2[i].value = cell1               # 赋值到test单元格
    wb2.save(excelpath2)                 # 保存数据
    wb1.close()                          # 关闭excel
    wb2.close()


#openpyxl写入数据
class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet
    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)


if __name__ == '__main__':
    a=Webrequests()
    a.test0001()
    # a.readSheetdata("J2")

    # filepath = "D:/auto_test_nancy/dfw_union/data/explore_union_case.xlsx"
    # sheetName = "Sheet1"
    # data = ExcelUtil(filepath,sheetName)
    # print(data.dict_data())

    #copy_excel("D:/auto_test_nancy/dfw_union/data/explore_union_case.xlsx", "D:/auto_test_nancy/dfw_union/data/explore_union_case_copy.xlsx")
    # wt = Write_excel("D:/auto_test_nancy/dfw_union/data/explore_union_case_copy.xlsx")
    # wt.write(4, 5, "HELLEOP")
    # wt.write(4, 6, "HELLEOP")