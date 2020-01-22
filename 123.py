# # -*- coding:UTF-8 -*-
#一些demo 练习
'''
# from public.common.log import Log
# import smtplib
# from email.mime.text import MIMEText
# from email.header import Header
#
# logger = Log()
# def send_email(sender='lusainan13579@163.com',login_pwd='150061803670513',
#                receivers=['lusainan13579@163.com']
#                ,subject='Python SMTP 邮件测试',
#                msg='Python 邮件发送测试...'):
#     """发送邮件"""
#     message = MIMEText(msg, 'plain', 'utf-8')
#     message['From'] = Header("菜鸟教程", 'utf-8')   # 发送者
#     message['To'] =  Header("测试", 'utf-8')        # 接收者
#     message['Subject'] = Header(subject, 'utf-8')
#     try:
#         smtpObj = smtplib.SMTP_SSL('smtp.163.com') #SMTP服务器
#         smtpObj.login(sender,login_pwd)  #不是登录邮箱密码，而是授权码
#         smtpObj.sendmail(sender, receivers, message.as_string())
#         smtpObj.quit()
#         logger.info ("邮件发送成功")
#     except smtplib.SMTPException:
#         logger.info ("Error: 无法发送邮件")
#
#
# if __name__=='__main__':
#     send_email()
#     pass



from json.encoder import JSONEncoder

from openpyxl import load_workbook
from macpath import split
import re
from urllib3.fields import RequestField
from urllib3.filepost import encode_multipart_formdata
from coingame.config.globalparam import get_testEv_ip_name_pwd
from public.common.log import Log
from decimal import Decimal

import paramiko
import os,sys
import smtplib
from email.mime.text import MIMEText
from email.header import Header
import datetime

#
# logger = Log()
# isNeedOne=False
# arr=['caseInfo','isRun=true','1','Y']
# # arr=['caseInfo','isRun=true','0']
# # print(len(arr))  #3
# # print(arr[2]) #0
# if(len(arr)>3):
#     needField=arr[3]           #needField=自定义的arr[3]  第四个字段
#     print(needField)   #Y
# isNeedOne=True if arr[2]=='1' else False  #if arr[2]=='1', isNeedOne=True else  False
# print(isNeedOne)

#excel编辑保存

class OperatorExcel(object):
    """数据库类"""
    #openpyxl也只能处理xlsx、xlsm文件，对xls也只能改xlwt 、wlrd这个两个模块处理！！！！！！！
    def __init__(self, filePath):    #构造方法
        self.filePath=filePath
        self.wb = load_workbook(filePath)  #openpyxl只支持xlsx格式，老版的xls格式需要其他方法去加载。
        # #获取全部表名
        # print("sheetnames",self.wb.sheetnames)
        #获取第一个表名
        # print("sheetname1:",self.wb.sheetnames[0])
        #建新表，放最后
        # ws1 = self.wb.create_sheet("Mysheet")  #NG
        # print(ws1)
        #修改表的名称


    def close(self):
        self.wb.close()

    def __parseSql(self,sql):  ##私有变量  __XX私有变量、_X实例变量、__XX__特殊变量
        arr=sql.split('&')  #将字符串$sql 按照 & 分割
        self.sheetName=arr[0]      #caseInfo&isRun=true&0  caseInfo表名 len(arr)==3
        self.where = arr[1].split(",");
        # print(self.where)
        if(len(arr)>3):
            self.needField=arr[3]           #needField=自定义的arr[3]  sql变量第四个字段,&后
        self.isNeedOne=True if arr[2]=='1' else False  ##if arr[2]=='1', isNeedOne=True else  False


    def getData2DictList(self,sql):
        """
         * 该方法支持多条件查询 sheetName&fieldName1=findValue,fieldName2=findValue2,...
         * (或者fieldName<>findValue,...)&inNeedOne
        """
        """1、用法一 读取匹配的行+列，组成[{},{},{}]嵌套，方便后面读取"""
        logger.info('执行的查询语句：'+sql)
        result = []
        rows = self.__matchCols(sql)  #调用下面__matchCols函数
        # print("1",rows)  #匹配isRun=true的行[2,3,6]
        sheet = self.wb[self.sheetName]
        # print("2",sheet)#<Worksheet "caseInfo">
        colNum = sheet.max_column
        # print("3",colNum)  #列max
        for i in rows:  #匹配几行循环几次  i 行
            temp = {}
            for m in range(1,colNum+1):  #1,2,3,4,5,6,7,8  循环8次 [1,colNum+1)   m列
                value = sheet.cell(row=i,column=m).value
                if value==None:
                    value=''
                temp[sheet.cell(row=1,column=m).value]=value  #temp['key']=value  字典赋值
            # print(temp)  #每个匹配的行组成一个字典
            result.append(temp)  #将字典拼接成list   [{},{},{}]
        self.__clearSql()          #初始化值清空为''
        # print(result)   #拼接好的list  [{},{},{},{}]
        return result

    def getOneField2List(self,sql):
        """
         * 该方法支持多条件查询 sheetName&fieldName1=findValue,fieldName2=findValue2,...
         * (或者fieldName<>findValue,...)&inNeedOne&needField
         * 注:needField为要读取的字段名称!!
        """
        """2、用法二 读取匹配范围内某字段值，组成list[]，方便后面读取"""
        result = []
        rows = self.__matchCols(sql) #匹配的行
        # print("match rows:",rows) #match rows: [2, 3, 6]
        sheet = self.wb.get_sheet_by_name(self.sheetName)
        fieldMap = self.__fieldAndIndex2Dict(self.sheetName)  #调用下面的私有函数__fieldAndIndex2Dict
        colNum = sheet.max_column
        # print("colNum:",colNum) #8列
        for i in rows:
            value = sheet.cell(row=i,column=fieldMap[self.needField]).value
            # print("value:",value)
            if value==None:
                value=''
#             result.append(sheet.cell(row=i,column=fieldMap[self.needField]).value)
            result.append(value)
        self.__clearSql()
        return result


    def insertValues(self,dict1,sheetName):
        #插入数据

        sheet = self.wb.get_sheet_by_name(sheetName)
        # print("sheet:",sheet)
        rowNum=sheet.max_row
        # print(rowNum)
        fieldMap = self.__fieldAndIndex2Dict(sheetName)
        # print("字段从1开始排序：",fieldMap)
        for d,v in dict1.items():
            try:
                sheet.cell(row=rowNum+1,column=fieldMap[d]).value=v
            except KeyError:
                pass
        self.__clearSql()
        # print("11",self.wb.read_only) #False
        self.__save()

    def updateValues(self,dict1,sql):
        #按sql 更新数据   比如这里替换2、3、6行数据
        """
         * 该方法支持多条件查询 sheetName&fieldName1=findValue,fieldName2=findValue2,...
         * (或者fieldName<>findValue,...)&inNeedOne
        """

        rows = self.__matchCols(sql)
        # print(rows)
        sheet = self.wb.get_sheet_by_name(self.sheetName)
        # print(sheet)
        rowNum=sheet.max_row
        # print(rowNum)
        fieldMap = self.__fieldAndIndex2Dict(self.sheetName)
        # print("字段从1开始排序:",fieldMap)  #fieldMap: {'id': 1, 'b_u_id': 2, 'user_name': 3, 'user_password': 4, 'betting_currency_id': 5, 'isRun': 6, 'caseInfo': 7, 'index': 8}
        for i in rows:
            for d,v in dict1.items():
                try:
                    sheet.cell(row=i,column=fieldMap[d]).value=v
                except KeyError:
                    pass
        self.__clearSql()
        self.__save()

    def __save(self):
        print("self.filePath:",self.filePath)
        self.wb.__save(self.filePath)



    def __matchCols(self,sql):
        #匹配满足条件的行数
        self.__parseSql(sql) #调用上面的__parseSql(sql) #获取self.sheetName、self.where、self.isNeedOne
        sheet = self.wb[self.sheetName]
        fieldMap = self.__fieldAndIndex2Dict(self.sheetName)  #调用下面的__fieldAndIndex2Dict
        rowNum = sheet.max_row
        colNum = sheet.max_column
        result = []
        for i in range(2,rowNum+1):
            for j in range(len(self.where)):
                where= re.split('=|<>',self.where[j])
                findField = where[0]
                findValue = where[1]
                value = sheet.cell(row=i,column=fieldMap[findField]).value
                if not self.__matchFindType(self.where[j], value, findValue):
                    break
            #此处的else（如果前面语句没有发生break语句 则会执行else 否则不执行else）
            else:
                result.append(i)
                #如果只需要一条数据则返回 否则继续
                if self.isNeedOne:
                    break
        return result
        # 1 [2, 3，6] 满足sql查询条件的行数

    #怎么理解?
    def __matchFindType(self,where,haveFindValue,exceptValue):
        findType='equals'
        haveFindValue=str(haveFindValue)
        if where.__contains__('<>'):
            findType="except"
        if findType=='equals':
            if exceptValue==haveFindValue:
                return True
        elif findType=='except':
            if exceptValue != haveFindValue:
                return True
        return False



    #把第一行的值 和 列号存入字典中  key=field value=列号（从1开始）
    #{'id': 1, 'b_u_id': 2, 'user_name': 3, 'user_password': 4, 'betting_currency_id': 5, 'isRun': 6, 'caseInfo': 7, 'index': 8}
    def __fieldAndIndex2Dict(self,sheetName):
        sheet = self.wb[sheetName]
        colNum=sheet.max_column
        map = {}
        for i in range(colNum):
            key = sheet.cell(row=1,column=i+1).value
            map[key]=i+1
        return map

    def __clearSql(self):
        self.sheetName=''
        self.where=''
        self.isNeedOne=False
        self.needField=''



if __name__=='__main__':
    path = 'D:\\auto_test_nancy\\coingame\\data\\cases.xlsx'
    excel = OperatorExcel(path)
    # data = excel.getData2DictList('caseInfo&isRun=true&0')  #将表caseInfo，字段值isRun=true匹配的行列数据组成list，每一行是一个dict  [{},{},{}]
    # data=excel.getOneField2List('caseInfo&isRun=true&0&caseInfo') #将表caseInfo,字段值isRun=true匹配行,且将字段caseInfo的字段值取出，拼接成list  [a,b,c]  0所有匹配(非1)  1匹第一个  if arr[2]=='1', isNeedOne=True else  False
    data=excel.insertValues({'id':'1', 'b_u_id':'UB1','user_name':'AUTO08@qq.com'},"caseInfo")

    #
    # #将字典的 key 和 value 组成一个新的列表
    # d={1:"a",2:"b",3:"c"}
    # result=[]
    # for k,v in d.items():
    #     result.append(k)
    #     result.append(v)
    #
    # print(result) #[1, 'a', 2, 'b', 3, 'c']
'''
'''
from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook('D:\\auto_test_nancy\\coingame\\data\\cases.xlsx')
print("初始：",wb.get_sheet_names())  # 获得所有sheet的名称
sheet=wb.get_sheet_by_name('caseInfo')
# print(sheet) #根据表名获得名称
print(sheet.title)#获取表名
# # # print(wb.active)#获得当前正在显示的sheet
# wb.create_sheet(index=1, title="mysheet1")  ##新建sheet======
# print("新建后：",wb.get_sheet_names())  # 获得所有sheet的名称
# wb.remove(wb[wb.sheetnames[0]])      ##删除sheet======
# print("删除后:",wb.get_sheet_names())  # 获得所有sheet的名称


sheet1 = wb[wb.sheetnames[1]]  # 打开一个 sheet 工作表
# print(sheet1.max_column)

# b4 = sheet['B4']
# print(f'({b4.column}, {b4.row}) is {b4.value}')
#
b4_too = sheet.cell(row=4, column=2)  #行4 列2
print("4-2",b4_too.value)

# print(sheet.max_row)  #行
# print(sheet.max_column) #列

#写入数据

sheet1.cell(row=4, column=2, value="B1")  # 修改单元格的值UB3为B1
sheet1["A2"] = "1a"  # 直接修改A1单元格的值为A1
sheet1["B7"] = "B7"  # 新增B11单元格的值为B11
sheet1.title = "test_sheet"  # 修改sheet1的表名为test_sheet

print("最终：",wb.get_sheet_names())


wb.save('D:\\auto_test_nancy\\coingame\\data\\cases.xlsx')

'''
#!/usr/bin/env python
# -*- coding:utf-8 -*-
'''
from datetime import datetime,timedelta
import time
import datetime
def add_sub_date(old_date,day:int=None,date_format='%Y-%m-%d %H:%M:%S'):
    """加减日期"""
    final_date = None
    if day != None:
        delta=datetime.timedelta(days=day)
        # print("delta:",delta) #delta: 1 day, 0:00:00
        final_date = old_date+delta
        print("final_date:",final_date)
    return final_date.strftime(date_format)

def add_sec_date(old_date,seconds:int=None,date_format='%Y-%m-%d %H:%M:%S'):
    """加减秒"""
    final_date = None
    if seconds != None:
        delta=datetime.timedelta(seconds=seconds)
        final_date = old_date+delta
    return final_date.strftime(date_format)


if __name__=='__main__':
    now1 = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    print("now1:",now1)
    now=datetime.datetime.now()
    print("now:",now)
    add_sub_date(now,1)
'''


#
# import re
# content = 'Hello 123456789 Word_This is just a test 666 Test'
# result = re.search('(\d+).*?(\d+).*', content)
# # print(result)
# # print(result.group())    # print(result.group(0)) 同样效果字符串
# # print(result.groups())
# print(result.group(1))


# #coding=utf-8
#
# import requests
# from coingame.config.globalparam import get_project_path, get_testEv, get_testUrl
# from coingame.module.db import coinGameDb
# import re
# from coingame.beans import public_values as w_v
# import json
# import copy
#
#
# interface_data_path=get_project_path()+'/coingame/data/coingame/'
#
# class TestEnvironment():
#
#     def __init__(self,domain=None,testEv=None,isCrm=True,http='http',is_pre=True):
#         self.game_play_old_status = '' #玩法操作之前的状态
#         self.game_play_new_status = '' #玩法操作之后的状态
#
#         if domain == None:
#             self.testEv = get_testEv()
#             self.testUrl = get_testUrl()
#         else:
#             self.testEv=testEv
#             self.testUrl='%s://%s/%s'%(http,domain,testEv)
#         testEv = self.testEv
#
#         if is_pre:
#             self.testUrl = re.sub('/pre$','',self.testUrl)
#
#         print(self.testEv,self.testUrl)
#         self.invalid_league_file = interface_data_path+'invalid_league_ls.txt'
#         self.invalid_eventId_file = interface_data_path+'invalid_eventId_ls.txt'
#         print(self.invalid_league_file,self.invalid_eventId_file)
#
#         if self.testEv=='testing':
#             self.db=coinGameDb(host='172.17.1.128',db='sp_test')
#         elif self.testEv=='abtest':
#             self.db=coinGameDb(db='sp_abtest')
#         elif self.testEv=='demo':
#             self.db=coinGameDb(db='sp_demo')
#         elif self.testEv=='pz':
#             self.db=coinGameDb(db='sp_pz_test')
#         elif self.testEv=='beta':
#             self.db=coinGameDb(host='172.17.3.163',db='sp_beta')
#         elif self.testEv=='inte1' or self.testEv=='inte':
#             self.db=coinGameDb(db='sp_test',test_ev=self.testEv)
#         else:
#             self.db=coinGameDb(db='sp_test')
#
#         if isCrm:
#             self.Origin='%s://%s-crm.intranet.etcgame.com'%(http,testEv) #crm
#             if testEv == 'pre':
#                 self.Origin='%s://%s-crm.coingame.com'%(http,testEv)
#         else:
#             self.Origin='%s://%s-www.intranet.etcgame.com'%(http,testEv) #website
#             if testEv == 'pre':
#                 self.Origin='%s://%s-www.coingame.com'%(http,testEv)
#         print("self.Origin",self.Origin)
#         self.session = requests.Session() #利用session保持登陆
#         self.headers = {'Authorization':'Basic YnJvd3Nlcjo=','Content-Type':'application/x-www-form-urlencoded','Origin':self.Origin}
#         # self.headers = {'Authorization':'Basic YnJvd3Nlcjo=','Origin':self.Origin}
#
#     def clear_attr(self):
#         self.game_play_new_status = ''
#         self.game_play_new_status = ''
#         self.pre_run_step_name = '' #运行的前一个步骤
#
#
#     def public_send_request(self,request_name='',
#                 data=None,headers=None,params=None
#                 ,file_dict=None,content_type=None,is_encode=True):
#
#         if isinstance(params,dict):
#             params = json.dumps(params,ensure_ascii=False)
#
#         if headers == None:
#             headers = copy.deepcopy(self.headers)
#
#
#         if content_type != None:
#             headers['Content-Type'] = content_type
#
#         request_arr = request_name.split(',')
#         url = self.testUrl + request_arr[0]
#         request_type = request_arr[1]
#
#         if file_dict != None :
#             # headers['Content-Type'] = w_v.upimage_content_type
#             headers.pop('Content-Type')
#             return self.session.post(url,files=file_dict,headers=headers)
#
#
#         # request_params_dict = {'url':url,'headers':headers,'files':file_dict}
#         request_params_dict = {'url':url,'headers':headers}
#         if is_encode and params != None:
#             params = params.encode()
#         if params != None:
#             request_params_dict.setdefault('params',params)
#         if data != None:
#             try :
#                 request_params_dict.setdefault('data',data.encode())
#             except:
#                 request_params_dict.setdefault('data',data)
#
#         if request_type == 'post':
#             return self.session.post(**request_params_dict)
#         if request_type == 'get':
#             # print('url=',url)
#             return self.session.get(**request_params_dict)
#         if request_type == 'put':
#
#             return self.session.put(**request_params_dict)
#         if request_type == 'options':
#             return self.session.options(**request_params_dict)
#
# if __name__=='__main__':
#     #t=TestEnvironment(domain=None,testEv=None,isCrm=True,http='http',is_pre=True)
#     t=TestEnvironment()  #全部默认

# tup1 = ('Google', 'Runoob', 1997, 2000)
# print(tup1[0])
#
# tup2 =(('Google','facebook','twitter'), 'Runoob', 1997, 2000)
# print(tup2[0][1])


import re
# p = re.compile('error(?!abc)')
# print("123")
# print(p.search("abc"))

# source_str="博风米饭-蛋烧鸡"
# reg_str="博风米饭"
# result = re.compile(reg_str,source_str)
# result.group()
# ret = re.match("博风米饭","博风米饭-蛋烧鸡")
# print(ret.group())
#
# ret = re.match("嫦娥1号","嫦娥1号发射成功")
# print(ret.group())
#
# ret = re.match("嫦娥\d号","嫦娥1号发射成功")
# print(ret.group())
#
ret = re.match("博风米饭","博风米饭-蛋烧鸡")
# if ret.group()=='博风米饭':
# if ret:
#     print("1")
# else:
#     print("2")
#
#
# ret = re.match("博风米饭","面食-蛋烧鸡")
# # if ret.group()=='博风米饭':
# if ret:
#     print("1")
# else:
#     print("2")

#
#
# a = input("input:")
#
# print(a)

import random

# randomStr= ""

# for i in range(6):
#     temp = random.randrange(0,3)
#     if temp == 0:
#         ch = chr(random.randrange(ord('A'),ord('Z') + 1))
#         randomStr += ch
#     elif temp == 1:
#         ch = chr(random.randrange(ord('a'),ord('z') + 1))
#         randomStr += ch
#     else:
#         ch = str((random.randrange(0,10)))
#         randomStr += ch
# print(randomStr)
#

# for i in range(6):
#     temp = random.randrange(0,2)
#     if temp == 0:
#         ch = chr(random.randrange(ord('a'),ord('z') + 1))
#         randomStr += ch
#     else:
#         ch = str((random.randrange(0,10)))
#         randomStr += ch
# print(randomStr)

# lists=['陆赛男1','陆赛男2','陆赛男3']
# for i in lists:
#     print(i)

#1到100求和

def qiuhe1(start,end):
    sum = 0
    for n in range(start,end+1,1):
        sum = sum + n
    return sum
print(qiuhe1(1,100))

def qiuhe2():
    sum = 0
    for n in range(1,101):
        sum = sum + n
    return sum
print(qiuhe2())

def qiuhe3():
    sum = 0
    x=1
    while x < 101:
        sum = sum + x
        x+=1
    return sum
print(qiuhe3())

from functools import reduce
def qiuhe4(x, y):
    return x + y
print(reduce(qiuhe4, range(1,101)))
print("end"+"\n")
